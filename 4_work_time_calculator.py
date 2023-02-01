import math

from openpyxl import load_workbook

"""
To Do:
- on start, show current monthly and quarterly data, and wether or not full-time covers quarterly limit
- reminder to fill in missing hours according to today. (do you want to fill it now?)
- monthly comparison to quarterly limit time (show how much of the quarterly time was used, how many overtime hours can be worked)
- future calculation for wether or not the employee will be able to work additional time (or if ful-time will cover the rest of the quarterly time)
- Google Calendar
- calculating monthly full-time working hours 9to substract form quarterly limit):  how many days are in the month (filter FRi_Sat), calculat sun-wed 9 hours, thu 8.5 hours, substract holidays.
- ability to see current data without entering the excel file.

Bugs:
- crashes if there's only clock in time (and 'clock out' is empty).
"""


if __name__ == '__main__':

    def open_xlsx_file():
        excel_path = "Assets\Book.xlsx"
        wb = load_workbook(filename=excel_path)
        ws = wb['Sheet1']  # !!! wb['Dummy'] #
        return ws

    def get_month_row_range(month):
        working_days_year = (25 * 12) + 1
        months_column = 1
        month_row_range = None;
        # get month's row range:
        for i in range(1, working_days_year):
            # read from the right month
            if ws.cell(row=i, column=months_column).value == month:
                month_row_range = [i + 2, i + 25]
        # Check if month exists in file
        if not month_row_range:
            return None
        else:
            return month_row_range

    def calculate_monthly(month):
        full_day_minutes = 9 * 60
        clock_in_column = 2
        clock_out_column = 3
        month_row_range = get_month_row_range(month)
        monthly_debt_time_minutes, monthly_spare_time_minutes, monthly_worked_minutes = 0, 0, 0
        # retrieve clocking in and out data:
        for i in range(month_row_range[0], month_row_range[1]):
            user_data = {
                "clocked_in": str(ws.cell(i, clock_in_column).value),
                "clocked_out": str(ws.cell(i, clock_out_column).value)
            }
            # break at the end of the month (clock-in cell value is None)
            if user_data["clocked_in"] == 'None':
                break;
            # skip special ocations:
            elif not ":" in user_data["clocked_in"]:
                print(f"Skipping this day because it says: {user_data['clocked_in']} \n===== ")
                continue;

            daily_time_worked_minutes = calculate_daily_work_time(user_data["clocked_in"], user_data["clocked_out"])
            # Total time worked this month
            monthly_worked_minutes += daily_time_worked_minutes
            print(f"You worked {seperate_hours_minutes(daily_time_worked_minutes)[0]} hours and {seperate_hours_minutes(daily_time_worked_minutes)[1]} minutes")
            if daily_time_worked_minutes < full_day_minutes:
                daily_debt_time_minutes = calculate_debt_time(daily_time_worked_minutes, full_day_minutes)
                print(f"You are in DEBT of {seperate_hours_minutes(daily_debt_time_minutes)[0]} hours and {seperate_hours_minutes(daily_debt_time_minutes)[1]} minutes \n=====")
                monthly_debt_time_minutes += daily_debt_time_minutes
            else:
                daily_spare_time_minutes = calculate_spare_time(daily_time_worked_minutes, full_day_minutes)
                print(f"You have {seperate_hours_minutes(daily_spare_time_minutes)[0]} hours and {seperate_hours_minutes(daily_spare_time_minutes)[1]} minutes to SPARE \n=====")
                monthly_spare_time_minutes += daily_spare_time_minutes

        monthly_tuple = 0
        if monthly_spare_time_minutes > monthly_debt_time_minutes:
            monthly_tuple = seperate_hours_minutes(monthly_spare_time_minutes - monthly_debt_time_minutes)
            print(f">>>Your MONTHLY SPARE time is: {monthly_tuple[0]} hours and {monthly_tuple[1]} minutes.")
            print(f" === END OF DATA FOR MONTH {month} ===\n")
            return "spare", combine_hours_minutes(monthly_tuple[0], monthly_tuple[1]), monthly_worked_minutes
        elif monthly_debt_time_minutes > monthly_spare_time_minutes:
            monthly_tuple = seperate_hours_minutes(monthly_debt_time_minutes - monthly_spare_time_minutes)
            print(f">>>Your MONTHLY DEBT time is:{monthly_tuple[0]} hours and {monthly_tuple[1]} minutes")
            print(f" === END OF DATA FOR MONTH {month} ===\n")
            return "debt", combine_hours_minutes(monthly_tuple[0], monthly_tuple[1]), monthly_worked_minutes
        else:
            print(">>>Your MONTHLY spare and debt times are the same.")
            print(f" === END OF DATA FOR MONTH {month} ===\n")
            return "same", combine_hours_minutes(monthly_tuple[0], monthly_tuple[1]), monthly_worked_minutes

    def calculate_quarterly(quarterly_months_list, quarterly_limit_minutes):
        quarterly_total_spare, quarterly_total_debt, quarterly_worked_minutes = 0, 0, 0
        for month in quarterly_months_list:
            # Check if month exists in file
            month_row_range = get_month_row_range(month)
            if not month_row_range:
                print(f">>> Skipping {month} because it's not in the file")
                continue
            monthly_summary = calculate_monthly(month)
            # Total time worked this quarterly
            quarterly_worked_minutes += monthly_summary[2]

            # calculate quarterly spare
            if monthly_summary[0] == "spare":
                quarterly_total_spare += monthly_summary[1]
            else:
                quarterly_total_debt += monthly_summary[1]

        # print quarterly summary
        if quarterly_total_spare > quarterly_total_debt:
            quarterly_spare_time_tuple = seperate_hours_minutes(quarterly_total_spare - quarterly_total_debt)
            print(f"\n>>>Your QUARTERLY SPARE time is: {quarterly_spare_time_tuple[0]} hours and {quarterly_spare_time_tuple[1]} minutes.")
        elif quarterly_total_debt > quarterly_total_spare:
            quarterly_debt_time_tuple = seperate_hours_minutes(quarterly_total_debt - quarterly_total_spare)
            print(f"\n>>>Your QUARTERLY DEBT time is:{quarterly_debt_time_tuple[0]} hours and {quarterly_debt_time_tuple[1]} minutes")
        else:
            print("\n>>>Your QUARTERLY spare and debt times are the same.")

        # print total quarterly working time
        if quarterly_limit_minutes and quarterly_worked_minutes > quarterly_limit_minutes:
            print(f"\033[1;31m \n>>>This QUARTERLY you worked {seperate_hours_minutes(quarterly_worked_minutes)[0]} hours and {seperate_hours_minutes(quarterly_worked_minutes)[1]} minutes")
            print(f"\033[1;31m \n>>>!!! Your quarterly working time is passing the quarterly working time limit of {seperate_hours_minutes(quarterly_limit_minutes)[0]} hours !!!")
        else:
            print(f"\033[1;32m \n>>>This QUARTERLY you worked {seperate_hours_minutes(quarterly_worked_minutes)[0]} hours and {seperate_hours_minutes(quarterly_worked_minutes)[1]} minutes")




    def get_quarterly(quarters, month):
        """
        :param month: month from user's input.
        :return: the quarterly the month is in.
        """
        return [k for k, v in quarters.items() if month in v][0]


    def calculate_daily_work_time(time_clock_in, time_clock_out):
        """
        :param time_clock_in: What time the employee started working
        :param time_clock_out: What time the employee finished working
        :return: the amount of time the employee working in hours and minutes
        """
        hours_clock_in, minutes_clock_in = int(time_clock_in.split(':')[0]), int(time_clock_in.split(':')[1])
        hours_clock_out, minutes_clock_out = int(time_clock_out.split(':')[0]), int(time_clock_out.split(':')[1])

        if minutes_clock_out < minutes_clock_in:
            minutes_worked = 60 - (minutes_clock_in - minutes_clock_out)
            hours_worked = hours_clock_out - hours_clock_in - 1
        else:
            minutes_worked = minutes_clock_out - minutes_clock_in
            hours_worked = hours_clock_out - hours_clock_in

        return combine_hours_minutes(hours_worked,minutes_worked)


    def calculate_debt_time(time_worked_minutes, goal_minutes):
        return goal_minutes - time_worked_minutes

    def calculate_spare_time(time_worked_minutes, goal_minutes):
        return time_worked_minutes - goal_minutes

    def combine_hours_minutes(hours,minutes):
        """
        :param hours: hours
        :param minutes: minutes
        :return: total time in minutes (hours + minutes combined)
        """
        full_time = (hours * 60) + minutes
        return full_time

    def seperate_hours_minutes(total_time):
        """
        :param total_time: total time in minutes
        :return: total_time seperated to hours and minutes
        """
        hours = math.floor(total_time / 60)
        minutes = total_time % 60
        return hours,minutes




def main():
    quarters = {
        "q1": ["JANUARY", "FEBRUARY", "MARCH"],
        "q2": ["APRIL", "MAY", "JUNE"],
        "q3": ["JULY", "AUGUST", "SEPTEMBER"],
        "q4": ["OCTOBER", "NOVEMBER", "DECEMBER"]
    }
    quarterly_limit_minutes = 546 * 60
    global ws
    ws = open_xlsx_file()

    # Getting month from user
    while True:
        month = input("Which MONTH do you want to retrieve data for? (full month name): ").upper()
        month_row_range = get_month_row_range(month)
        if not month_row_range:
            print(f"month {month} wasn't found in file. Please enter a different month. \n")
        else:
            # month is in file, continue with the program
            break

    # getting function from user
    while True:
        function = input("What output would you like? (Monthly/Quarterly): ").upper()
        print("\n")
        if function == "MONTHLY":
            calculate_monthly(month)
            break
        elif function == "QUARTERLY":
            calculate_quarterly(quarters[get_quarterly(quarters, month)], quarterly_limit_minutes)
            break
        else:
            print(f"The function {function} doesn't exist. Please choose one of the existing functions. \n")


main()